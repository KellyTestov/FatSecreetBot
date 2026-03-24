"""
Экспорт данных в Excel и PDF.

Excel: полноценный отчёт с несколькими листами.
PDF: отчёт с таблицами, поддержка кириллицы через системный шрифт Arial (Windows).
"""
import os
import platform
from pathlib import Path
from datetime import date

import config
from analytics import period_summary, top_products, parse_entry, to_float
from formatters import fmt_date, fmt_num


def _ensure_exports_dir() -> Path:
    config.EXPORTS_DIR.mkdir(exist_ok=True)
    return config.EXPORTS_DIR


# ============================================================
# Excel экспорт
# ============================================================

def create_excel(days_data: dict, start: date, end: date, goals: dict) -> Path:
    """
    Создаёт Excel файл с тремя листами:
    - Сводка: итоги и средние за период
    - По дням: одна строка на день
    - Все записи: каждый продукт отдельной строкой
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Лист 1: Сводка ----
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    summary = period_summary(days_data)
    period_label = f"{fmt_date(start)} — {fmt_date(end)}"

    ws_summary["A1"] = f"Отчёт по питанию: {period_label}"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")

    row = 3
    ws_summary[f"A{row}"] = "Показатель"
    ws_summary[f"B{row}"] = "Значение"
    ws_summary[f"A{row}"].font = header_font
    ws_summary[f"B{row}"].font = header_font
    row += 1

    def add_row(label, value):
        nonlocal row
        ws_summary[f"A{row}"] = label
        ws_summary[f"B{row}"] = value
        row += 1

    add_row("Дней в периоде", summary["total_days"])
    add_row("Дней с записями", summary["days_with_data"])
    add_row("Дней без записей", summary["days_without_data"])

    if summary["totals"]:
        add_row("", "")
        add_row("--- Итого за период ---", "")
        add_row("Калории (ккал)", round(summary["totals"]["calories"], 1))
        add_row("Белки (г)", round(summary["totals"]["protein"], 1))
        add_row("Жиры (г)", round(summary["totals"]["fat"], 1))
        add_row("Углеводы (г)", round(summary["totals"]["carbs"], 1))
        add_row("Сахар (г)", round(summary["totals"]["sugar"], 1))
        add_row("Натрий (мг)", round(summary["totals"]["sodium"], 1))

        add_row("", "")
        add_row("--- Средние за день ---", "")
        add_row("Калории (ккал)", round(summary["averages"]["calories"], 1))
        add_row("Белки (г)", round(summary["averages"]["protein"], 1))
        add_row("Жиры (г)", round(summary["averages"]["fat"], 1))
        add_row("Углеводы (г)", round(summary["averages"]["carbs"], 1))

        if summary["best_day"]:
            add_row("", "")
            add_row("Лучший день (мин. ккал)", f"{fmt_date(summary['best_day'])} ({fmt_num(summary['best_day_calories'])} ккал)")
            add_row("Худший день (макс. ккал)", f"{fmt_date(summary['worst_day'])} ({fmt_num(summary['worst_day_calories'])} ккал)")

    if any(v is not None for v in goals.values()):
        add_row("", "")
        add_row("--- Цели ---", "")
        if goals.get("calories") is not None:
            add_row("Цель калории", goals["calories"])
        if goals.get("protein") is not None:
            add_row("Цель белок (мин)", goals["protein"])
        if goals.get("max_sugar") is not None:
            add_row("Максимум сахара", goals["max_sugar"])
        if goals.get("max_sodium") is not None:
            add_row("Максимум натрия", goals["max_sodium"])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # ---- Лист 2: По дням ----
    ws_days = wb.create_sheet("По дням")
    day_headers = ["Дата", "Калории", "Белки (г)", "Жиры (г)", "Углеводы (г)", "Сахар (г)", "Натрий (мг)"]
    ws_days.append(day_headers)

    # Стиль заголовков
    for col_idx, _ in enumerate(day_headers, 1):
        cell = ws_days.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if summary["daily_totals"]:
        for d in sorted(summary["daily_totals"].keys()):
            t = summary["daily_totals"][d]
            ws_days.append([
                fmt_date(d),
                round(t["calories"], 1),
                round(t["protein"], 1),
                round(t["fat"], 1),
                round(t["carbs"], 1),
                round(t["sugar"], 1),
                round(t["sodium"], 1),
            ])

    for col_idx in range(1, len(day_headers) + 1):
        ws_days.column_dimensions[get_column_letter(col_idx)].width = 14

    # ---- Лист 3: Все записи ----
    ws_entries = wb.create_sheet("Все записи")
    entry_headers = ["Дата", "Приём пищи", "Продукт", "Описание", "Калории", "Белки (г)", "Жиры (г)", "Углеводы (г)", "Сахар (г)", "Натрий (мг)"]
    ws_entries.append(entry_headers)

    for col_idx, _ in enumerate(entry_headers, 1):
        cell = ws_entries.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    meal_names = {
        "Breakfast": "Завтрак",
        "Lunch": "Обед",
        "Dinner": "Ужин",
        "Snacks": "Перекус",
        "Other": "Другое",
    }

    for d in sorted(days_data.keys()):
        for entry in days_data[d]:
            parsed = parse_entry(entry)
            ws_entries.append([
                fmt_date(d),
                meal_names.get(parsed["meal"], parsed["meal"]),
                parsed["name"],
                parsed["description"],
                round(parsed["calories"], 1),
                round(parsed["protein"], 1),
                round(parsed["fat"], 1),
                round(parsed["carbs"], 1),
                round(parsed["sugar"], 1),
                round(parsed["sodium"], 1),
            ])

    col_widths = [12, 12, 35, 25, 10, 10, 10, 12, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws_entries.column_dimensions[get_column_letter(col_idx)].width = width

    # Сохраняем файл
    exports_dir = _ensure_exports_dir()
    filename = f"nutrition_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    filepath = exports_dir / filename
    wb.save(filepath)
    return filepath


# ============================================================
# PDF экспорт
# ============================================================

def _register_cyrillic_font():
    """
    Регистрирует шрифт с поддержкой кириллицы для reportlab.
    На Windows использует системный Arial.
    Возвращает (обычный_шрифт, жирный_шрифт).
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # 1) Явные пути через env (удобно для Railway/контейнеров)
    env_regular = os.getenv("PDF_FONT_REGULAR")
    env_bold = os.getenv("PDF_FONT_BOLD")
    if env_regular and os.path.exists(env_regular):
        pdfmetrics.registerFont(TTFont("CyrRegular", env_regular))
        if env_bold and os.path.exists(env_bold):
            pdfmetrics.registerFont(TTFont("CyrBold", env_bold))
        else:
            pdfmetrics.registerFont(TTFont("CyrBold", env_regular))
        return "CyrRegular", "CyrBold"

    # 2) Локальный fonts/ в репозитории (если захотим положить TTF рядом с кодом)
    local_candidates = [
        (Path("fonts/DejaVuSans.ttf"), Path("fonts/DejaVuSans-Bold.ttf")),
        (Path("fonts/Arial.ttf"), Path("fonts/Arial-Bold.ttf")),
    ]
    for regular_path, bold_path in local_candidates:
        if regular_path.exists():
            pdfmetrics.registerFont(TTFont("CyrRegular", str(regular_path)))
            if bold_path.exists():
                pdfmetrics.registerFont(TTFont("CyrBold", str(bold_path)))
            else:
                pdfmetrics.registerFont(TTFont("CyrBold", str(regular_path)))
            return "CyrRegular", "CyrBold"

    # 3) Системные шрифты для разных ОС
    if platform.system() == "Windows":
        font_paths = [
            ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
            ("C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
        ]
        for regular, bold in font_paths:
            if os.path.exists(regular) and os.path.exists(bold):
                pdfmetrics.registerFont(TTFont("CyrRegular", regular))
                pdfmetrics.registerFont(TTFont("CyrBold", bold))
                return "CyrRegular", "CyrBold"
    elif platform.system() == "Linux":
        linux_paths = [
            ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            ("/usr/share/fonts/dejavu/DejaVuSans.ttf", "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
            ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
        ]
        for regular, bold in linux_paths:
            if os.path.exists(regular):
                pdfmetrics.registerFont(TTFont("CyrRegular", regular))
                if os.path.exists(bold):
                    pdfmetrics.registerFont(TTFont("CyrBold", bold))
                else:
                    pdfmetrics.registerFont(TTFont("CyrBold", regular))
                return "CyrRegular", "CyrBold"
    elif platform.system() == "Darwin":
        mac_paths = [
            ("/System/Library/Fonts/Supplemental/Arial Unicode.ttf", "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
            ("/System/Library/Fonts/Supplemental/Arial.ttf", "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ]
        for regular, bold in mac_paths:
            if os.path.exists(regular):
                pdfmetrics.registerFont(TTFont("CyrRegular", regular))
                if os.path.exists(bold):
                    pdfmetrics.registerFont(TTFont("CyrBold", bold))
                else:
                    pdfmetrics.registerFont(TTFont("CyrBold", regular))
                return "CyrRegular", "CyrBold"

    # Fallback: Helvetica (кириллица не отобразится, но PDF создастся)
    return "Helvetica", "Helvetica-Bold"


def create_pdf(days_data: dict, start: date, end: date, goals: dict) -> Path:
    """
    Создаёт PDF отчёт с:
    - заголовком периода
    - сводной таблицей
    - таблицей по дням
    - топ продуктов
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    font_regular, font_bold = _register_cyrillic_font()

    exports_dir = _ensure_exports_dir()
    filename = f"nutrition_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
    filepath = exports_dir / filename

    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    style_title = ParagraphStyle("title", fontName=font_bold, fontSize=16, alignment=TA_CENTER, spaceAfter=12)
    style_h2 = ParagraphStyle("h2", fontName=font_bold, fontSize=12, spaceBefore=12, spaceAfter=6)
    style_normal = ParagraphStyle("normal", fontName=font_regular, fontSize=10, spaceAfter=4)

    header_bg = colors.HexColor("#4472C4")
    row_alt = colors.HexColor("#EEF2FF")

    def make_table(data, col_widths=None):
        t = Table(data, colWidths=col_widths)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 1), (-1, -1), font_regular),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, row_alt]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
        t.setStyle(style)
        return t

    summary = period_summary(days_data)
    period_label = f"{fmt_date(start)} — {fmt_date(end)}"

    story = []

    # Заголовок
    story.append(Paragraph(f"Отчёт по питанию", style_title))
    story.append(Paragraph(period_label, style_title))
    story.append(Spacer(1, 0.5 * cm))

    # Сводная таблица
    story.append(Paragraph("Сводка за период", style_h2))
    summary_data = [["Показатель", "Значение"]]
    summary_data.append(["Дней в периоде", str(summary["total_days"])])
    summary_data.append(["Дней с записями", str(summary["days_with_data"])])
    summary_data.append(["Дней без записей", str(summary["days_without_data"])])

    if summary["totals"]:
        summary_data.append(["Калории итого (ккал)", fmt_num(summary["totals"]["calories"])])
        summary_data.append(["Белки итого (г)", fmt_num(summary["totals"]["protein"])])
        summary_data.append(["Жиры итого (г)", fmt_num(summary["totals"]["fat"])])
        summary_data.append(["Углеводы итого (г)", fmt_num(summary["totals"]["carbs"])])
        summary_data.append(["Калории средние (ккал)", fmt_num(summary["averages"]["calories"])])
        summary_data.append(["Белки средние (г)", fmt_num(summary["averages"]["protein"])])
        summary_data.append(["Жиры средние (г)", fmt_num(summary["averages"]["fat"])])
        summary_data.append(["Углеводы средние (г)", fmt_num(summary["averages"]["carbs"])])
        if summary["best_day"]:
            summary_data.append(["Лучший день (мин. ккал)", f"{fmt_date(summary['best_day'])} ({fmt_num(summary['best_day_calories'])}ккал)"])
            summary_data.append(["Худший день (макс. ккал)", f"{fmt_date(summary['worst_day'])} ({fmt_num(summary['worst_day_calories'])}ккал)"])

    story.append(make_table(summary_data, col_widths=[10 * cm, 6 * cm]))
    story.append(Spacer(1, 0.5 * cm))

    # По дням
    if summary["daily_totals"]:
        story.append(Paragraph("По дням", style_h2))
        days_table_data = [["Дата", "Калории", "Белки г", "Жиры г", "Углеводы г"]]
        for d in sorted(summary["daily_totals"].keys()):
            t = summary["daily_totals"][d]
            days_table_data.append([
                fmt_date(d),
                fmt_num(t["calories"]),
                fmt_num(t["protein"]),
                fmt_num(t["fat"]),
                fmt_num(t["carbs"]),
            ])
        story.append(make_table(days_table_data, col_widths=[4 * cm, 3.5 * cm, 3 * cm, 3 * cm, 3.5 * cm]))
        story.append(Spacer(1, 0.5 * cm))

    # Топ продуктов
    products = top_products(days_data, n=10, sort_by="calories")
    if products:
        story.append(Paragraph("Топ продуктов по калориям", style_h2))
        prod_data = [["Продукт", "Раз", "Калории", "Белки г"]]
        for p in products:
            prod_data.append([
                p["name"][:45],
                str(p["count"]),
                fmt_num(p["calories"]),
                fmt_num(p["protein"]),
            ])
        story.append(make_table(prod_data, col_widths=[9 * cm, 2 * cm, 3 * cm, 3 * cm]))

    doc.build(story)
    return filepath


def _status_float(row: dict, key: str) -> float | None:
    raw = (row.get(key) or "").strip().replace(",", ".")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _status_yes(row: dict, key: str) -> bool:
    return (row.get(key) or "").strip().lower() in {"да", "yes", "y", "1", "true"}


def _metric_zone_name(metric: str, value: float) -> str:
    if metric == "protein":
        if value < 100:
            return "red"
        if value < 120:
            return "yellow"
        return "green"
    if metric == "carbs":
        if value < 90 or value > 210:
            return "red"
        if value <= 160:
            return "green"
        return "yellow"
    if metric == "fiber":
        if value < 10:
            return "red"
        if value < 18:
            return "yellow"
        return "green"
    if metric == "fat":
        if value < 40 or value > 80:
            return "red"
        if value <= 65:
            return "green"
        return "yellow"
    return "green"


def create_weekly_pdf_report(
    *,
    week_num: int,
    start: date,
    end: date,
    days_data: dict,
    status_rows: list[dict],
    goals: dict,
    start_date: date,
) -> Path:
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font_regular, font_bold = _register_cyrillic_font()
    exports_dir = _ensure_exports_dir()
    filename = f"weekly_week_{week_num}_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
    filepath = exports_dir / filename

    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
    )
    style_title = ParagraphStyle("title", fontName=font_bold, fontSize=15, alignment=TA_CENTER, spaceAfter=8)
    style_h2 = ParagraphStyle("h2", fontName=font_bold, fontSize=11, spaceBefore=8, spaceAfter=4)
    style_p = ParagraphStyle("p", fontName=font_regular, fontSize=9, alignment=TA_LEFT, leading=12, spaceAfter=3)

    header_bg = colors.HexColor("#264653")
    row_alt = colors.HexColor("#F3F6FA")

    summary = period_summary(days_data)
    daily = summary.get("daily_totals", {})
    ordered_days = sorted(daily.keys())

    status_by_date = {}
    for row in status_rows:
        d = row.get("_date")
        if d:
            status_by_date[d] = row

    zone_counts = {"red": 0, "yellow": 0, "green": 0}
    for row in status_rows:
        for header, metric in (("Белок (г)", "protein"), ("Углеводы (г)", "carbs"), ("Клетчатка (г)", "fiber"), ("Жиры (г)", "fat")):
            v = _status_float(row, header)
            if v is None:
                continue
            zone_counts[_metric_zone_name(metric, v)] += 1

    activity_yes = {
        "Зал": sum(1 for row in status_rows if _status_yes(row, "Зал")),
        "Кардио": sum(1 for row in status_rows if _status_yes(row, "Кардио")),
        "Бассейн": sum(1 for row in status_rows if _status_yes(row, "Бассейн")),
    }

    weights = []
    for row in status_rows:
        w = _status_float(row, "Вес (кг)")
        if w is not None:
            weights.append((row["_date"], w))
    weights.sort(key=lambda x: x[0])
    week_start_weight = weights[0][1] if weights else None
    week_end_weight = weights[-1][1] if weights else None
    week_delta = (week_end_weight - week_start_weight) if (week_start_weight is not None and week_end_weight is not None) else None

    start_weight = None
    for row in sorted(status_rows, key=lambda x: x.get("_date", "")):
        w = _status_float(row, "Вес (кг)")
        if w is not None:
            start_weight = w
            break
    if start_weight is None and weights:
        start_weight = weights[0][1]
    total_delta = (week_end_weight - start_weight) if (start_weight is not None and week_end_weight is not None) else None

    def make_table(data, col_widths=None):
        t = Table(data, colWidths=col_widths)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), header_bg),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), font_bold),
                    ("FONTNAME", (0, 1), (-1, -1), font_regular),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.4),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, row_alt]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        return t

    def bar_daily_calories() -> Drawing:
        draw = Drawing(520, 180)
        chart = VerticalBarChart()
        chart.x = 45
        chart.y = 35
        chart.height = 120
        chart.width = 450
        vals = [daily[d]["calories"] for d in ordered_days] if ordered_days else [0]
        chart.data = [vals]
        chart.barWidth = 12
        chart.groupSpacing = 8
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = max(1000, int(max(vals) * 1.2) if vals else 1000)
        chart.valueAxis.valueStep = max(100, int(chart.valueAxis.valueMax / 6))
        chart.categoryAxis.labels.boxAnchor = "ne"
        chart.categoryAxis.labels.angle = 30
        chart.categoryAxis.categoryNames = [d.strftime("%d.%m") for d in ordered_days] if ordered_days else ["-"]
        chart.bars[0].fillColor = colors.HexColor("#2A9D8F")
        draw.add(chart)
        draw.add(String(8, 164, "Калории по дням", fontName=font_bold, fontSize=10))
        return draw

    def bar_daily_macros() -> Drawing:
        draw = Drawing(520, 190)
        chart = VerticalBarChart()
        chart.x = 45
        chart.y = 35
        chart.height = 130
        chart.width = 450
        protein = [daily[d]["protein"] for d in ordered_days] if ordered_days else [0]
        carbs = [daily[d]["carbs"] for d in ordered_days] if ordered_days else [0]
        fat = [daily[d]["fat"] for d in ordered_days] if ordered_days else [0]
        chart.data = [protein, carbs, fat]
        chart.barWidth = 4
        chart.groupSpacing = 8
        chart.barSpacing = 1
        chart.valueAxis.valueMin = 0
        mx = max(protein + carbs + fat) if ordered_days else 100
        chart.valueAxis.valueMax = max(100, int(mx * 1.25))
        chart.valueAxis.valueStep = max(10, int(chart.valueAxis.valueMax / 6))
        chart.categoryAxis.labels.boxAnchor = "ne"
        chart.categoryAxis.labels.angle = 30
        chart.categoryAxis.categoryNames = [d.strftime("%d.%m") for d in ordered_days] if ordered_days else ["-"]
        chart.bars[0].fillColor = colors.HexColor("#1D3557")
        chart.bars[1].fillColor = colors.HexColor("#457B9D")
        chart.bars[2].fillColor = colors.HexColor("#E9C46A")
        draw.add(chart)
        draw.add(String(8, 172, "Макросы по дням (г)", fontName=font_bold, fontSize=10))
        draw.add(String(350, 172, "Белок", fontName=font_regular, fontSize=8, fillColor=colors.HexColor("#1D3557")))
        draw.add(String(390, 172, "Углеводы", fontName=font_regular, fontSize=8, fillColor=colors.HexColor("#457B9D")))
        draw.add(String(455, 172, "Жиры", fontName=font_regular, fontSize=8, fillColor=colors.HexColor("#E9C46A")))
        return draw

    def pie_zones() -> Drawing:
        draw = Drawing(250, 180)
        pie = Pie()
        pie.x = 45
        pie.y = 15
        pie.width = 145
        pie.height = 145
        pie.data = [zone_counts["red"], zone_counts["yellow"], zone_counts["green"]]
        pie.labels = [f"Красн. {zone_counts['red']}", f"Жёлт. {zone_counts['yellow']}", f"Зел. {zone_counts['green']}"]
        pie.slices[0].fillColor = colors.HexColor("#E76F51")
        pie.slices[1].fillColor = colors.HexColor("#E9C46A")
        pie.slices[2].fillColor = colors.HexColor("#2A9D8F")
        draw.add(pie)
        draw.add(String(5, 166, "Зоны нутриентов", fontName=font_bold, fontSize=10))
        return draw

    def pie_activity() -> Drawing:
        draw = Drawing(250, 180)
        pie = Pie()
        pie.x = 45
        pie.y = 15
        pie.width = 145
        pie.height = 145
        yes_total = sum(activity_yes.values())
        no_total = max(0, summary["total_days"] * 3 - yes_total)
        pie.data = [yes_total, no_total]
        pie.labels = [f"Сделано {yes_total}", f"Пропуск {no_total}"]
        pie.slices[0].fillColor = colors.HexColor("#2A9D8F")
        pie.slices[1].fillColor = colors.HexColor("#BFC8D6")
        draw.add(pie)
        draw.add(String(5, 166, "Активность (зал/кардио/бассейн)", fontName=font_bold, fontSize=9))
        return draw

    def line_weight() -> Drawing:
        draw = Drawing(520, 180)
        chart = HorizontalLineChart()
        chart.x = 45
        chart.y = 35
        chart.height = 120
        chart.width = 450
        data = [w for _, w in weights]
        if not data:
            data = [0]
        chart.data = [data]
        chart.valueAxis.valueMin = min(data) - 1 if len(data) > 1 else max(0, data[0] - 1)
        chart.valueAxis.valueMax = max(data) + 1 if len(data) > 1 else data[0] + 1
        chart.valueAxis.valueStep = max(0.2, (chart.valueAxis.valueMax - chart.valueAxis.valueMin) / 6)
        chart.categoryAxis.categoryNames = [datetime.fromisoformat(d).strftime("%d.%m") for d, _ in weights] if weights else ["-"]
        chart.lines[0].strokeColor = colors.HexColor("#E76F51")
        chart.lines[0].strokeWidth = 2
        draw.add(chart)
        draw.add(String(8, 164, "Тренд веса", fontName=font_bold, fontSize=10))
        return draw

    story = []
    story.append(Paragraph("Недельный отчет похудения", style_title))
    story.append(Paragraph(f"Неделя {week_num}: {fmt_date(start)} - {fmt_date(end)}", style_title))
    story.append(
        Paragraph(
            f"Старт похудения: {fmt_date(start_date)}. Дней с данными: {summary['days_with_data']} из {summary['total_days']}.",
            style_p,
        )
    )

    totals = summary["totals"] or {"calories": 0.0, "protein": 0.0, "carbs": 0.0, "fat": 0.0, "fiber": 0.0}
    avgs = summary["averages"] or {"calories": 0.0, "protein": 0.0, "carbs": 0.0, "fat": 0.0}
    kpi = [
        ["Показатель", "Значение"],
        ["Калории всего", f"{totals['calories']:.0f} ккал"],
        ["Калории среднее/день", f"{avgs['calories']:.1f} ккал"],
        ["Белок среднее/день", f"{avgs['protein']:.1f} г"],
        ["Углеводы среднее/день", f"{avgs['carbs']:.1f} г"],
        ["Жиры среднее/день", f"{avgs['fat']:.1f} г"],
        ["Зал/Кардио/Бассейн (Да)", f"{activity_yes['Зал']} / {activity_yes['Кардио']} / {activity_yes['Бассейн']}"],
        ["Вес старт недели", "-" if week_start_weight is None else f"{week_start_weight:.1f} кг"],
        ["Вес конец недели", "-" if week_end_weight is None else f"{week_end_weight:.1f} кг"],
        ["Изм. за неделю", "-" if week_delta is None else f"{week_delta:+.1f} кг"],
        ["Изм. от старта", "-" if total_delta is None else f"{total_delta:+.1f} кг"],
    ]
    story.append(Spacer(1, 0.2 * cm))
    story.append(make_table(kpi, col_widths=[7.8 * cm, 8.2 * cm]))
    story.append(Spacer(1, 0.25 * cm))

    if any(v is not None for v in goals.values()):
        story.append(Paragraph("План vs факт", style_h2))
        goals_data = [["Метрика", "Факт", "Цель", "Отклонение"]]
        if goals.get("calories") is not None:
            diff = avgs["calories"] - goals["calories"]
            goals_data.append(["Калории", f"{avgs['calories']:.1f}", f"{goals['calories']}", f"{diff:+.1f}"])
        if goals.get("protein") is not None:
            diff = avgs["protein"] - goals["protein"]
            goals_data.append(["Белок", f"{avgs['protein']:.1f}", f"{goals['protein']}", f"{diff:+.1f}"])
        if len(goals_data) > 1:
            story.append(make_table(goals_data, col_widths=[5 * cm, 3.5 * cm, 3.5 * cm, 4 * cm]))
            story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("Графики и диаграммы", style_h2))
    story.append(bar_daily_calories())
    story.append(Spacer(1, 0.2 * cm))
    story.append(bar_daily_macros())
    story.append(Spacer(1, 0.2 * cm))
    story.append(line_weight())
    story.append(Spacer(1, 0.1 * cm))
    story.append(Table([[pie_zones(), pie_activity()]], colWidths=[8.2 * cm, 8.2 * cm]))

    if summary["daily_totals"]:
        story.append(Paragraph("Дневная таблица", style_h2))
        daily_data = [["Дата", "Ккал", "Белок", "Углев.", "Клетч.", "Жиры", "Вес", "Аппетит", "Тирз", "Зал", "Кардио", "Бассейн"]]
        for d in sorted(summary["daily_totals"].keys()):
            t = summary["daily_totals"][d]
            key = d.isoformat()
            row = status_by_date.get(key, {})
            daily_data.append(
                [
                    fmt_date(d),
                    f"{t['calories']:.0f}",
                    f"{t['protein']:.1f}",
                    f"{t['carbs']:.1f}",
                    f"{t['fiber']:.1f}",
                    f"{t['fat']:.1f}",
                    row.get("Вес (кг)", ""),
                    (row.get("Аппетит / ощущения", "") or "")[:28],
                    row.get("Тирзетта", ""),
                    row.get("Зал", ""),
                    row.get("Кардио", ""),
                    row.get("Бассейн", ""),
                ]
            )
        story.append(make_table(daily_data, col_widths=[1.7 * cm, 1.2 * cm, 1.1 * cm, 1.1 * cm, 1.1 * cm, 1.1 * cm, 1.1 * cm, 3 * cm, 1.5 * cm, 0.8 * cm, 0.9 * cm, 1.1 * cm]))

    products = top_products(days_data, n=10, sort_by="calories")
    if products:
        story.append(Paragraph("Топ продуктов недели", style_h2))
        prod_data = [["Продукт", "Раз", "Ккал", "Белок", "Углев.", "Жиры"]]
        for p in products:
            prod_data.append([p["name"][:34], str(p["count"]), f"{p['calories']:.0f}", f"{p['protein']:.1f}", f"{p['carbs']:.1f}", f"{p['fat']:.1f}"])
        story.append(make_table(prod_data, col_widths=[8.4 * cm, 1.2 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm]))

    recommendations = []
    if avgs["protein"] < 120:
        recommendations.append("Подними средний белок: целись минимум в 120 г/день.")
    if avgs["carbs"] > 210 or avgs["carbs"] < 90:
        recommendations.append("Стабилизируй углеводы в коридоре 90-160 г.")
    if avgs["fat"] > 80 or avgs["fat"] < 40:
        recommendations.append("Верни жиры в коридор 40-65 г.")
    if week_delta is not None:
        recommendations.append(f"Вес за неделю: {week_delta:+.1f} кг.")
    if not recommendations:
        recommendations.append("Неделя по макрорамкам выглядит стабильно, удерживай текущую структуру.")

    story.append(Paragraph("Краткий вывод", style_h2))
    for line in recommendations:
        story.append(Paragraph(f"- {line}", style_p))

    doc.build(story)
    return filepath
